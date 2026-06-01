from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, Response
import schemas, auth, database
from typing import List, Optional, Any
import razorpay
from urllib.parse import quote
from config import settings
from datetime import datetime, timedelta, timezone, date
import calendar
import io

# openpyxl for Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

router = APIRouter(prefix="/api", tags=["Orders"])


def strip_oid(doc):
    """Recursively remove MongoDB _id fields."""
    if isinstance(doc, list):
        return [strip_oid(d) for d in doc]
    if isinstance(doc, dict):
        return {k: strip_oid(v) for k, v in doc.items() if k != "_id"}
    return doc

razorpay_client = None
if settings.RAZORPAY_KEY_ID and settings.RAZORPAY_KEY_SECRET:
    razorpay_client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))

@router.post("/orders", response_model=schemas.OrderResponse)
async def create_order(order_data: schemas.OrderCreate, db: Any = Depends(database.get_db), current_user: Optional[Any] = Depends(auth.get_current_user_optional)):
    total = 0
    items_to_create = []
    
    for item in order_data.items:
        product = await db.products.find_one({"id": item.product_id})
        if not product:
            raise HTTPException(status_code=404, detail=f"Product {item.product_id} not found")
        
        # Handle color variant stock
        if item.color_variant_id:
            variant = await db.product_color_variants.find_one({
                "id": item.color_variant_id,
                "product_id": item.product_id
            })
            if not variant:
                raise HTTPException(status_code=404, detail=f"Color variant not found")
            if variant["stock"] < item.quantity:
                raise HTTPException(status_code=400, detail=f"Insufficient stock for {product['name']} - {variant['color_name']}")
            
            # Update variant stock
            await db.product_color_variants.update_one({"id": item.color_variant_id}, {"$inc": {"stock": -item.quantity}})
        else:
            # Use product-level stock
            if product["stock"] < item.quantity:
                raise HTTPException(status_code=400, detail=f"Insufficient stock for {product['name']}")
            await db.products.update_one({"id": item.product_id}, {"$inc": {"stock": -item.quantity}})
        
        price = product["discounted_price"] if product.get("discounted_price") else product["original_price"]
        total += price * item.quantity
        
        items_to_create.append({
            "product_id": product["id"],
            "color_variant_id": item.color_variant_id,
            "product_name": product["name"],
            "selected_color": item.selected_color,
            "selected_size": item.selected_size,
            "price": price,
            "quantity": item.quantity
        })

    # Calculate shipping cost
    shipping_cost = 0
    if total < 700 and total > 0:
        shipping_cost = 50
    elif total >= 700 and total <= 1200:
        shipping_cost = 30
    
    grand_total = total + shipping_cost

    # Combine address fields
    full_address = order_data.shipping_address
    if order_data.shipping_city: full_address += f", {order_data.shipping_city}"
    if order_data.shipping_state: full_address += f", {order_data.shipping_state}"
    if order_data.shipping_pincode: full_address += f" - {order_data.shipping_pincode}"

    # Generate Order ID
    count_doc = await db.counters.find_one({"_id": "orders"})
    if not count_doc:
        max_order = await db.orders.find_one(sort=[("id", -1)])
        start_val = max_order.get("id", 0) if max_order else 0
        await db.counters.update_one({"_id": "orders"}, {"$set": {"seq": start_val}}, upsert=True)
    
    new_id = await database.get_next_id(db, "orders")

    new_order = {
        "id": new_id,
        "user_id": current_user.id if current_user else None,
        "customer_name": order_data.customer_name,
        "customer_email": order_data.customer_email,
        "customer_phone": order_data.customer_phone,
        "address": full_address,
        "total_amount": grand_total,
        "status": "pending",
        "payment_method": order_data.payment_method,
        "created_at": datetime.utcnow()
    }
    
    await db.orders.insert_one(new_order)
    
    # Save order items
    for item in items_to_create:
        # Initialize order_items counter
        count_doc = await db.counters.find_one({"_id": "order_items"})
        if not count_doc:
            max_item = await db.order_items.find_one(sort=[("id", -1)])
            start_val = max_item.get("id", 0) if max_item else 0
            await db.counters.update_one({"_id": "order_items"}, {"$set": {"seq": start_val}}, upsert=True)
            
        item_id = await database.get_next_id(db, "order_items")
        item["id"] = item_id
        item["order_id"] = new_id
        await db.order_items.insert_one(item)
    
    new_order["items"] = items_to_create
    return strip_oid(new_order)

@router.get("/orders", response_model=List[schemas.OrderResponse])
async def get_user_orders(current_user: Any = Depends(auth.get_current_user), db: Any = Depends(database.get_db)):
    orders = await db.orders.find({"user_id": current_user.id}).sort("created_at", -1).to_list(length=100)
    for order in orders:
        order["items"] = await db.order_items.find({"order_id": order["id"]}).to_list(length=100)
    return strip_oid(orders)

@router.get("/orders/{order_id}", response_model=schemas.OrderResponse)
async def get_order(order_id: int, current_user: Any = Depends(auth.get_current_user), db: Any = Depends(database.get_db)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    if order["user_id"] != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    order["items"] = await db.order_items.find({"order_id": order_id}).to_list(length=100)
    return strip_oid(order)

@router.post("/orders/razorpay/create", response_model=schemas.RazorpayOrderResponse)
async def create_razorpay_payment(data: schemas.RazorpayOrderCreate, db: Any = Depends(database.get_db)):
    if not razorpay_client:
        raise HTTPException(status_code=500, detail="Razorpay not configured")
    order = await db.orders.find_one({"id": data.order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    try:
        rzp_order = razorpay_client.order.create({"amount": int(data.amount * 100), "currency": "INR", "receipt": f"order_{order['id']}"})
        await db.orders.update_one({"id": data.order_id}, {"$set": {"razorpay_order_id": rzp_order["id"]}})
        return {"razorpay_order_id": rzp_order["id"], "amount": data.amount, "currency": "INR", "key_id": settings.RAZORPAY_KEY_ID}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/orders/razorpay/verify")
async def verify_payment(data: schemas.RazorpayPaymentVerify, db: Any = Depends(database.get_db)):
    order = await db.orders.find_one({"id": data.order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    params_dict = {'razorpay_order_id': data.razorpay_order_id, 'razorpay_payment_id': data.razorpay_payment_id, 'razorpay_signature': data.razorpay_signature}
    try:
        razorpay_client.utility.verify_payment_signature(params_dict)
        await db.orders.update_one({"id": data.order_id}, {"$set": {"status": "confirmed", "razorpay_payment_id": data.razorpay_payment_id}})
        return {"message": "Success"}
    except Exception:
        raise HTTPException(status_code=400, detail="Payment verification failed")

@router.post("/orders/whatsapp", response_model=schemas.WhatsAppOrderResponse)
async def create_whatsapp_order(order_data: schemas.WhatsAppOrderCreate, db: Any = Depends(database.get_db), current_user: Optional[Any] = Depends(auth.get_current_user_optional)):
    standard_data = schemas.OrderCreate(
        customer_name=order_data.customer_name, 
        customer_email=order_data.customer_email,
        customer_phone=order_data.customer_phone, 
        shipping_address=order_data.shipping_address,
        shipping_city=order_data.shipping_city,
        shipping_state=order_data.shipping_state,
        shipping_pincode=order_data.shipping_pincode,
        payment_method="whatsapp", 
        items=order_data.items
    )
    order = await create_order(standard_data, db, current_user)
    
    msg = f"🛍️ *New Order #{order['id']}*\n\n"
    msg += f"👤 *Customer:* {order['customer_name']}\n"
    msg += f"📞 *Phone:* {order['customer_phone']}\n"
    msg += f"📍 *Shipping Address:* {order['address']}\n\n"
    msg += f"💰 *Total Amount:* ₹{order['total_amount']}\n\n"
    msg += f"📦 *Items:*\n"
    
    for item in order["items"]:
        details = []
        if item.get("selected_color"): details.append(f"Color: {item['selected_color']}")
        if item.get("selected_size"): details.append(f"Size: {item['selected_size']}")
        detail_str = f" ({', '.join(details)})" if details else ""
        
        msg += f"• {item['product_name']} x {item['quantity']}{detail_str}\n"
    
    whatsapp_url = f"https://wa.me/{settings.WHATSAPP_BUSINESS_NUMBER}?text={quote(msg)}"
    return {"order_id": order["id"], "whatsapp_url": whatsapp_url, "message": msg}

@router.get("/admin/orders", response_model=schemas.OrderListResponse)
async def get_all_orders_admin(
    page: int = 1,
    page_size: int = 20,
    status: Optional[str] = None,
    search: Optional[str] = None,
    admin: Any = Depends(auth.get_admin), 
    db: Any = Depends(database.get_db)
):
    import math
    filt = {}
    if status:
        filt["status"] = status
    if search:
        if search.isdigit():
            filt["$or"] = [
                {"id": int(search)},
                {"customer_phone": {"$regex": search, "$options": "i"}}
            ]
        else:
            filt["$or"] = [
                {"customer_name": {"$regex": search, "$options": "i"}},
                {"customer_email": {"$regex": search, "$options": "i"}},
                {"customer_phone": {"$regex": search, "$options": "i"}}
            ]
            
    total = await db.orders.count_documents(filt)
    pages = math.ceil(total / page_size) if page_size > 0 else 1
    offset = (page - 1) * page_size
    
    orders = await db.orders.find(filt).sort("created_at", -1).skip(offset).limit(page_size).to_list(length=page_size)
    for order in orders:
        order["items"] = await db.order_items.find({"order_id": order["id"]}).to_list(length=100)
        
    return {
        "items": strip_oid(orders),
        "total": total,
        "page": page,
        "page_size": page_size,
        "pages": pages
    }

@router.put("/admin/orders/{order_id}/status")
async def update_order_status(order_id: int, status_update: dict, admin: Any = Depends(auth.get_admin), db: Any = Depends(database.get_db)):
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="Order not found")
    
    new_status = status_update.get("status", order["status"])
    await db.orders.update_one({"id": order_id}, {"$set": {"status": new_status}})
    return {"message": "Status updated"}


# ---------------------------------------------------------------------------
# Helper: compute date range from filter name
# ---------------------------------------------------------------------------
def _get_date_range(filter_by: str, start_date: str = None, end_date: str = None):
    """Return (start_dt, end_dt) UTC-aware datetimes or (None, None) for 'all'."""
    now = datetime.now(timezone.utc)
    today = now.date()

    if filter_by == "all":
        return None, None

    elif filter_by == "current_year":
        start = datetime(today.year, 1, 1, tzinfo=timezone.utc)
        end = datetime(today.year, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "last_year":
        y = today.year - 1
        start = datetime(y, 1, 1, tzinfo=timezone.utc)
        end = datetime(y, 12, 31, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "current_month":
        start = datetime(today.year, today.month, 1, tzinfo=timezone.utc)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end = datetime(today.year, today.month, last_day, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "last_month":
        first_of_this = date(today.year, today.month, 1)
        last_month_end = first_of_this - timedelta(days=1)
        start = datetime(last_month_end.year, last_month_end.month, 1, tzinfo=timezone.utc)
        last_day = calendar.monthrange(last_month_end.year, last_month_end.month)[1]
        end = datetime(last_month_end.year, last_month_end.month, last_day, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "this_week":
        # ISO: Monday is start of week
        monday = today - timedelta(days=today.weekday())
        sunday = monday + timedelta(days=6)
        start = datetime(monday.year, monday.month, monday.day, tzinfo=timezone.utc)
        end = datetime(sunday.year, sunday.month, sunday.day, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "last_week":
        monday_this = today - timedelta(days=today.weekday())
        sunday_last = monday_this - timedelta(days=1)
        monday_last = sunday_last - timedelta(days=6)
        start = datetime(monday_last.year, monday_last.month, monday_last.day, tzinfo=timezone.utc)
        end = datetime(sunday_last.year, sunday_last.month, sunday_last.day, 23, 59, 59, tzinfo=timezone.utc)
        return start, end

    elif filter_by == "custom":
        if not start_date or not end_date:
            raise HTTPException(status_code=400, detail="start_date and end_date required for custom range")
        try:
            start = datetime.strptime(start_date, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            end = datetime.strptime(end_date, "%Y-%m-%d").replace(hour=23, minute=59, second=59, tzinfo=timezone.utc)
        except ValueError:
            raise HTTPException(status_code=400, detail="Dates must be in YYYY-MM-DD format")
        return start, end

    return None, None


# ---------------------------------------------------------------------------
# Helper: dynamic file name
# ---------------------------------------------------------------------------
def _get_filename(filter_by: str, start_date: str = None, end_date: str = None) -> str:
    mapping = {
        "all": "orders_all",
        "current_year": "orders_current_year",
        "last_year": "orders_last_year",
        "current_month": "orders_current_month",
        "last_month": "orders_last_month",
        "this_week": "orders_this_week",
        "last_week": "orders_last_week",
    }
    if filter_by == "custom" and start_date and end_date:
        return f"orders_custom_{start_date}_to_{end_date}"
    return mapping.get(filter_by, "orders_export")


# ---------------------------------------------------------------------------
# Export orders endpoint
# ---------------------------------------------------------------------------
@router.get("/admin/orders/export")
async def export_orders(
    filter_by: str = Query("all"),
    sort_by: str = Query("newest"),
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
    export_format: str = Query("xlsx"),
    token: Optional[str] = Query(None, description="JWT token for direct browser downloads"),
    db: Any = Depends(database.get_db),
):
    # Validate admin from query token (direct browser download)
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        from jose import jwt as _jwt, JWTError
        from config import settings as _settings
        payload = _jwt.decode(token, _settings.JWT_SECRET_KEY, algorithms=[_settings.JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
        user_id = payload.get("sub")
        user = await db.users.find_one({"id": int(user_id)})
        if not user or user.get("role") != "admin":
            raise HTTPException(status_code=403, detail="Admin access required")
    except Exception as e:
        raise HTTPException(status_code=401, detail=f"Invalid token: {e}")
    if openpyxl is None and export_format == "xlsx":
        raise HTTPException(status_code=500, detail="openpyxl is not installed on the server")

    # --- Build date filter ---
    start_dt, end_dt = _get_date_range(filter_by, start_date, end_date)
    query: dict = {}
    if start_dt and end_dt:
        query["created_at"] = {"$gte": start_dt, "$lte": end_dt}

    # --- Build sort ---
    sort_map = {
        "newest": [("created_at", -1)],
        "oldest": [("created_at", 1)],
        "amount_high": [("total_amount", -1)],
        "amount_low": [("total_amount", 1)],
    }
    sort_spec = sort_map.get(sort_by, [("created_at", -1)])

    # --- Fetch all orders (date filtering done in Python since created_at may be stored as string) ---
    sort_spec_mongo = [("created_at", -1)] if sort_by in ("newest", "oldest") else sort_map.get(sort_by, [("created_at", -1)])
    orders = await db.orders.find({}).sort(sort_spec_mongo).to_list(length=None)
    for order in orders:
        order["items"] = await db.order_items.find({"order_id": order["id"]}).to_list(length=None)

    # --- Filter by date in Python (handles both string and datetime created_at) ---
    def _parse_created_at(val):
        if isinstance(val, datetime):
            return val.replace(tzinfo=timezone.utc) if val.tzinfo is None else val
        if isinstance(val, str) and val:
            for fmt in ("%Y-%m-%d %H:%M:%S.%f", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
                try:
                    return datetime.strptime(val, fmt).replace(tzinfo=timezone.utc)
                except ValueError:
                    continue
        return None

    if start_dt and end_dt:
        filtered = []
        for order in orders:
            dt = _parse_created_at(order.get("created_at"))
            if dt and start_dt <= dt <= end_dt:
                filtered.append(order)
        orders = filtered

    # --- Sort in Python for amount-based sorts (string dates may not sort correctly in Mongo) ---
    if sort_by == "oldest":
        orders.sort(key=lambda o: (_parse_created_at(o.get("created_at")) or datetime.min.replace(tzinfo=timezone.utc)))
    elif sort_by == "amount_high":
        orders.sort(key=lambda o: o.get("total_amount", 0), reverse=True)
    elif sort_by == "amount_low":
        orders.sort(key=lambda o: o.get("total_amount", 0))

    # --- Prepare flat rows ---
    rows = []
    for order in orders:
        order_date_raw = order.get("created_at")
        dt = _parse_created_at(order_date_raw)
        if dt:
            order_date_str = dt.strftime("%Y-%m-%d %H:%M:%S")
        else:
            order_date_str = str(order_date_raw) if order_date_raw else ""

        payment_method = order.get("payment_method", "N/A")
        payment_id = order.get("razorpay_payment_id", "")
        order_status = order.get("status", "pending")
        items = order.get("items", [])
        
        base_row = {
            "Order ID": order.get("id", ""),
            "Customer Name": order.get("customer_name", ""),
            "Email": order.get("customer_email", ""),
            "Phone": order.get("customer_phone", ""),
            "Address": order.get("address", ""),
            "Order Total": order.get("total_amount", 0),
            "Payment Method": payment_method,
            "Payment ID": payment_id,
            "Order Status": order_status,
            "Order Date": order_date_str,
        }

        if not items:
            row = base_row.copy()
            row.update({
                "Product Name": "(no items)",
                "Color": "",
                "Size": "",
                "Quantity": 0,
                "Item Price": 0,
            })
            rows.append(row)
        else:
            for item in items:
                row = base_row.copy()
                row.update({
                    "Product Name": item.get("product_name", ""),
                    "Color": item.get("selected_color", ""),
                    "Size": item.get("selected_size", ""),
                    "Quantity": item.get("quantity", 0),
                    "Item Price": item.get("price", 0),
                })
                rows.append(row)

    filename_base = _get_filename(filter_by, start_date, end_date)
    columns_order = [
        "Order ID", "Order Date", "Order Status", "Customer Name", "Email", "Phone", "Address",
        "Product Name", "Color", "Size", "Quantity", "Item Price", "Order Total", 
        "Payment Method", "Payment ID"
    ]

    # ----------------------------------------------------------------
    # CSV export (fast path)
    # ----------------------------------------------------------------
    if export_format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=columns_order)
        writer.writeheader()
        writer.writerows(rows)
        csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
        return Response(
            content=csv_bytes,
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename_base}.csv"'},
        )

    # ----------------------------------------------------------------
    # XLSX export
    # ----------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders Data"

    # --- Theme colours ---
    HEADER_BG = "C62828"
    HEADER_FG = "FFFFFF"
    BORDER_CLR = "E0E0E0"

    thin = Side(style="thin", color=BORDER_CLR)
    full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # --- Write Header Row ---
    HEADER_ROW = 1
    for col_idx, col_name in enumerate(columns_order, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color=HEADER_FG)
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = full_border
        
        # Approximate column widths
        width = 15
        if col_name in ["Address", "Product Name", "Email"]: width = 30
        elif col_name in ["Order Date", "Payment ID"]: width = 20
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        
    ws.row_dimensions[HEADER_ROW].height = 20
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=1)

    # --- Data rows ---
    for row_idx, row_data in enumerate(rows, start=HEADER_ROW + 1):
        for col_idx, col_name in enumerate(columns_order, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data[col_name])
            cell.border = full_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # --- Auto-filter ---
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns_order))}{len(rows) + 1}"

    # --- Stream response ---
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename_base}.xlsx"',
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
