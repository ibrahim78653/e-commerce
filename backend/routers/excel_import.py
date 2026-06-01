"""
Excel Import/Export Router for Burhani Collection Admin Portal
Handles:
 - POST /api/admin/products/import-excel
 - GET  /api/admin/products/export-excel
 - GET  /api/admin/products/excel-template
 - GET  /api/admin/products/import-logs
"""

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, BackgroundTasks, Query
from fastapi.responses import StreamingResponse
from typing import Any, List, Optional
import io
import re
import uuid
from datetime import datetime
import auth
import database

# openpyxl for Excel read/write
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise RuntimeError("openpyxl is required. Run: pip install openpyxl")

router = APIRouter(prefix="/api/admin/products", tags=["Admin Excel"])

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
REQUIRED_COLUMNS = [
    "product_id", "product_name", "original_price",
    "discounted_price", "description", "color", "size",
    "product_images", "is_active", "stock", "category"
]
VALID_CATEGORIES = ["Ladies Wear", "Gents Wear", "Kids Wear", "Accessories", "Common"]
MAX_FILE_SIZE = 5 * 1024 * 1024  # 5 MB
MAX_DESCRIPTION_WORDS = 100


# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────

def word_count(text: str) -> int:
    return len(text.split()) if text else 0


def parse_bool(val) -> bool:
    """Parse TRUE/FALSE/1/0/yes/no strings robustly."""
    if isinstance(val, bool):
        return val
    if val is None or str(val).strip() == "":
        return True
    return str(val).strip().upper() in ("TRUE", "1", "YES", "Y")


def parse_csv(val) -> List[str]:
    """Split comma-separated string, strip whitespace, drop empties."""
    if not val:
        return []
    return [v.strip() for v in str(val).split(",") if v.strip()]


def is_valid_url(url: str) -> bool:
    pattern = re.compile(
        r'^(https?://)|(\/static\/)'
        r'|(^/[a-zA-Z0-9])',
        re.IGNORECASE
    )
    return bool(pattern.match(url.strip())) if url.strip() else True  # empty = ok


def sanitize(val) -> str:
    """Strip any leading '=' or formula injection characters from cell values."""
    s = str(val).strip() if val is not None else ""
    if s.startswith(("=", "+", "-", "@")):
        s = "'" + s  # prefix to neutralise formula injection
    return s


def validate_row(row_num: int, row: dict) -> List[str]:
    """Return list of error strings for a row. Empty list = row is valid."""
    errors = []

    pid = str(row.get("product_id", "")).strip()
    if not pid:
        errors.append(f"Row {row_num}: product_id is required.")

    try:
        op = float(row.get("original_price", 0) or 0)
        if op <= 0:
            errors.append(f"Row {row_num}: original_price must be > 0.")
    except (ValueError, TypeError):
        errors.append(f"Row {row_num}: original_price must be a number.")
        op = 0

    dp_raw = row.get("discounted_price")
    if dp_raw is not None and str(dp_raw).strip() not in ("", "None"):
        try:
            dp = float(dp_raw)
            if dp > op:
                errors.append(f"Row {row_num}: discounted_price ({dp}) must be ≤ original_price ({op}).")
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: discounted_price must be a number.")

    desc = str(row.get("description", "") or "")
    if word_count(desc) > MAX_DESCRIPTION_WORDS:
        errors.append(
            f"Row {row_num}: description exceeds {MAX_DESCRIPTION_WORDS} words "
            f"(got {word_count(desc)})."
        )

    stock_raw = row.get("stock")
    if stock_raw is not None and str(stock_raw).strip() not in ("", "None"):
        try:
            st = int(float(stock_raw))
            if st < 0:
                errors.append(f"Row {row_num}: stock must be >= 0.")
        except (ValueError, TypeError):
            errors.append(f"Row {row_num}: stock must be a whole number.")

    cat_raw = str(row.get("category", "") or "").strip()
    if cat_raw and cat_raw not in VALID_CATEGORIES:
        errors.append(f"Row {row_num}: category must be one of: {', '.join(VALID_CATEGORIES)}.")

    return errors


# ─────────────────────────────────────────────
#  BACKGROUND: perform the actual DB writes
# ─────────────────────────────────────────────

async def _get_next_product_id(db: Any) -> int:
    """Get next auto-increment ID for products collection."""
    count_doc = await db.counters.find_one({"_id": "products"})
    if not count_doc:
        max_prod = await db.products.find_one(sort=[("id", -1)])
        start_val = max_prod.get("id", 0) if max_prod else 0
        await db.counters.update_one({"_id": "products"}, {"$set": {"seq": start_val}}, upsert=True)
    res = await db.counters.find_one_and_update(
        {"_id": "products"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return res["seq"]


async def _get_next_image_id(db: Any) -> int:
    """Get next auto-increment ID for product_images collection."""
    res = await db.counters.find_one_and_update(
        {"_id": "product_images"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True
    )
    return res["seq"]


def _make_slug(name: str, suffix: str = "") -> str:
    """Generate a URL-friendly slug from a product name."""
    import re as _re
    slug = _re.sub(r'[^a-z0-9]+', '-', name.lower().strip()).strip('-')
    return f"{slug}-{suffix}" if suffix else slug


async def _sync_to_main_catalog(
    pid: str, name: str, op: float, dp, desc: str,
    colors: List[str], sizes: List[str], images: List[str],
    is_active: bool, stock: int, category: str, db: Any
) -> dict:
    """
    Upsert one product into the main db.products collection.
    Returns a dict with keys: action ('created'|'updated'|'already_present'), changed_fields.
    """
    existing = await db.products.find_one({"product_id": pid})
    colors_str = ",".join(colors)
    sizes_str = ",".join(sizes)

    if existing:
        # Detect changed fields
        changed_fields = []
        checks = [
            ("name",              existing.get("name"),              name),
            ("original_price",    existing.get("original_price"),    op),
            ("discounted_price",  existing.get("discounted_price"),  dp),
            ("description",       existing.get("description"),       desc),
            ("is_active",         existing.get("is_active"),         is_active),
            ("colors",            existing.get("colors", ""),         colors_str),
            ("sizes",             existing.get("sizes", ""),          sizes_str),
            ("stock",             existing.get("stock", 0),           stock),
            ("category",          existing.get("category", ""),       category),
        ]
        for field, old_val, new_val in checks:
            if str(old_val or "").strip() != str(new_val or "").strip():
                changed_fields.append(field)

        if not changed_fields and not images:
            return {"action": "already_present", "changed_fields": []}

        # Has changes — update
        update_data = {
            "name":             name,
            "original_price":   op,
            "discounted_price": dp,
            "description":      desc,
            "is_active":        is_active,
            "colors":           colors_str,
            "sizes":            sizes_str,
            "stock":            stock,
            "category":         category,
            "updated_at":       datetime.utcnow(),
        }
        await db.products.update_one({"product_id": pid}, {"$set": update_data})
        main_id = existing["id"]

        # Refresh images if provided
        if images:
            await db.product_images.delete_many({"product_id": main_id, "color_variant_id": None})
            for idx, url in enumerate(images):
                img_id = await _get_next_image_id(db)
                await db.product_images.insert_one({
                    "id": img_id, "product_id": main_id,
                    "color_variant_id": None, "image_url": url,
                    "is_primary": (idx == 0)
                })
            if "images" not in changed_fields:
                changed_fields.append("images")

        return {"action": "updated", "changed_fields": changed_fields}

    else:
        # New product — insert
        new_id = await _get_next_product_id(db)
        base_slug = _make_slug(name, str(new_id))
        # Ensure slug uniqueness
        existing_slug = await db.products.find_one({"slug": base_slug})
        slug = f"{base_slug}-{new_id}" if existing_slug else base_slug

        product_doc = {
            "id":               new_id,
            "product_id":       pid,
            "name":             name,
            "slug":             slug,
            "original_price":   op,
            "discounted_price": dp,
            "description":      desc,
            "is_active":        is_active,
            "colors":           colors_str,
            "sizes":            sizes_str,
            "stock":            stock,
            "category":         category,
            "category_id":      None,
            "is_featured":      False,
            "created_at":       datetime.utcnow(),
            "updated_at":       datetime.utcnow(),
        }
        await db.products.insert_one(product_doc)

        # Insert images
        for idx, url in enumerate(images):
            img_id = await _get_next_image_id(db)
            await db.product_images.insert_one({
                "id": img_id, "product_id": new_id,
                "color_variant_id": None, "image_url": url,
                "is_primary": (idx == 0)
            })

        return {"action": "created", "changed_fields": []}


async def _do_import(
    rows: List[dict],
    admin_id: int,
    partial: bool,
    db: Any,
    sections_uploaded: List[str],
):
    """
    Perform upsert logic for each validated row.
    Syncs to both excel_products and the main products catalog.
    Returns a rich categorised report.
    """
    # Report buckets
    newly_added: List[dict]     = []
    updated: List[dict]         = []
    already_present: List[dict] = []
    failed: List[dict]          = []

    for row in rows:
        pid = str(row["product_id"]).strip()
        name = sanitize(row.get("product_name", pid) or pid)
        try:
            colors   = parse_csv(row.get("color"))
            sizes    = parse_csv(row.get("size"))
            images   = parse_csv(row.get("product_images"))
            is_active = parse_bool(row.get("is_active", True))

            try:
                op = float(row.get("original_price", 0) or 0)
            except Exception:
                op = 0

            dp_raw = row.get("discounted_price")
            dp = None
            if dp_raw is not None and str(dp_raw).strip() not in ("", "None"):
                try:
                    dp = float(dp_raw)
                except Exception:
                    dp = None

            desc = sanitize(row.get("description", "") or "")

            stock_raw = row.get("stock")
            try:
                stock = int(float(stock_raw)) if stock_raw is not None and str(stock_raw).strip() not in ("", "None") else 0
            except Exception:
                stock = 0

            category = str(row.get("category", "") or "").strip()
            if category not in VALID_CATEGORIES:
                category = "Common"

            # ── 1. Upsert into excel_products (existing behaviour) ──
            excel_existing = await db.excel_products.find_one({"product_id": pid})
            if excel_existing:
                await db.excel_products.update_one(
                    {"product_id": pid},
                    {"$set": {
                        "product_name": name, "original_price": op,
                        "discounted_price": dp, "description": desc,
                        "is_active": is_active, "stock": stock,
                        "category": category, "updated_at": datetime.utcnow(),
                    }}
                )
                await db.excel_product_colors.delete_many({"product_id": pid})
                await db.excel_product_sizes.delete_many({"product_id": pid})
                await db.excel_product_images.delete_many({"product_id": pid})
            else:
                await db.excel_products.insert_one({
                    "product_id": pid, "product_name": name,
                    "original_price": op, "discounted_price": dp,
                    "description": desc, "is_active": is_active,
                    "stock": stock, "category": category,
                    "created_at": datetime.utcnow(), "updated_at": datetime.utcnow(),
                })

            for c in colors:
                await db.excel_product_colors.insert_one({"product_id": pid, "color_name": c})
            for s in sizes:
                await db.excel_product_sizes.insert_one({"product_id": pid, "size_value": s})
            for idx, img in enumerate(images):
                await db.excel_product_images.insert_one({"product_id": pid, "image_url": img, "sort_order": idx})

            # ── 2. Sync to main products catalog ──
            sync_result = await _sync_to_main_catalog(
                pid, name, op, dp, desc, colors, sizes, images, is_active, stock, category, db
            )
            action = sync_result["action"]
            changed_fields = sync_result["changed_fields"]

            entry = {"product_id": pid, "product_name": name}
            if action == "created":
                newly_added.append(entry)
            elif action == "updated":
                updated.append({**entry, "changed_fields": changed_fields})
            else:
                already_present.append(entry)

        except Exception as exc:
            failed.append({"product_id": pid, "product_name": name, "error": str(exc)})
            if not partial:
                await db.excel_import_logs.insert_one({
                    "log_id": str(uuid.uuid4()), "admin_id": admin_id,
                    "timestamp": datetime.utcnow(),
                    "newly_added_count": len(newly_added),
                    "updated_count": len(updated),
                    "already_present_count": len(already_present),
                    "failed_count": len(failed) + 1,
                    "status": "rolled_back",
                    "errors": failed,
                })
                raise

    # ── Write audit log ──
    await db.excel_import_logs.insert_one({
        "log_id":               str(uuid.uuid4()),
        "admin_id":             admin_id,
        "timestamp":            datetime.utcnow(),
        "newly_added_count":    len(newly_added),
        "updated_count":        len(updated),
        "already_present_count": len(already_present),
        "failed_count":         len(failed),
        "sections_uploaded":    sections_uploaded,
        "status":               "completed",
        "errors":               failed,
        # legacy fields for backward compat with old AuditLogs UI
        "created":              len(newly_added),
        "updated":              len(updated),
        "deactivated":          0,
        "failed":               len(failed),
    })

    return {
        "newly_added":          newly_added,
        "updated":              updated,
        "already_present":      already_present,
        "failed":               failed,
        "sections_uploaded":    sections_uploaded,
    }


# ─────────────────────────────────────────────
#  IMPORT ENDPOINT
# ─────────────────────────────────────────────

@router.post("/import-excel")
async def import_excel(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    partial: bool = Query(False, description="Allow partial success instead of full rollback"),
    admin: Any = Depends(auth.get_admin),
    db: Any = Depends(database.get_db),
):
    """
    Import products from an .xlsx file.
    Validates all rows BEFORE writing to DB.
    Syncs to main product catalog automatically.
    Returns a rich categorised import report.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted.")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File exceeds 5 MB limit.")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Could not read Excel file: {exc}")

    headers = [str(cell.value).strip().lower() if cell.value else "" for cell in ws[1]]
    sections_uploaded = [h for h in headers if h]  # non-empty column names

    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing_cols:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {', '.join(missing_cols)}"
        )

    rows: List[dict] = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        if all(c is None for c in row_cells):
            continue
        row_dict = {headers[i]: row_cells[i] for i in range(len(headers))}
        rows.append(row_dict)

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file contains no data rows.")

    seen_ids: set = set()
    dup_ids: List[str] = []
    for r in rows:
        pid = str(r.get("product_id", "")).strip()
        if pid in seen_ids:
            dup_ids.append(pid)
        seen_ids.add(pid)
    if dup_ids:
        raise HTTPException(
            status_code=422,
            detail=f"Duplicate product_id values found in file: {', '.join(set(dup_ids))}"
        )

    all_errors: List[dict] = []
    for idx, row in enumerate(rows, start=2):
        errs = validate_row(idx, row)
        if errs:
            all_errors.extend([{"row": idx, "product_id": str(row.get("product_id","")), "error": e} for e in errs])

    if all_errors and not partial:
        return {
            "status": "validation_failed",
            "message": "Fix the errors below and re-upload.",
            "errors": all_errors,
            "preview": _rows_to_preview(rows[:5]),
            "newly_added": [], "updated": [], "already_present": [], "failed": [],
            "sections_uploaded": sections_uploaded,
        }

    try:
        result = await _do_import(rows, admin.id, partial, db, sections_uploaded)
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Import failed: {exc}")

    return {
        "status": "completed",
        "total_rows": len(rows),
        "summary": {
            "newly_added_count":     len(result["newly_added"]),
            "updated_count":         len(result["updated"]),
            "already_present_count": len(result["already_present"]),
            "failed_count":          len(result["failed"]),
        },
        "newly_added":      result["newly_added"],
        "updated":          result["updated"],
        "already_present":  result["already_present"],
        "failed":           result["failed"],
        "sections_uploaded": result["sections_uploaded"],
        "preview":          _rows_to_preview(rows[:5]),
    }


def _rows_to_preview(rows: List[dict]) -> List[dict]:
    """Return cleaned preview rows for the frontend."""
    out = []
    for r in rows:
        stock_raw = r.get("stock")
        try:
            stock = int(float(stock_raw)) if stock_raw is not None and str(stock_raw).strip() not in ("", "None") else 0
        except Exception:
            stock = 0
        out.append({
            "product_id": sanitize(r.get("product_id", "")),
            "product_name": sanitize(r.get("product_name", "")),
            "original_price": r.get("original_price"),
            "discounted_price": r.get("discounted_price"),
            "color": r.get("color", ""),
            "size": r.get("size", ""),
            "is_active": parse_bool(r.get("is_active", True)),
            "stock": stock,
            "category": str(r.get("category", "") or "").strip(),
        })
    return out


# ─────────────────────────────────────────────
#  EXPORT ENDPOINT
# ─────────────────────────────────────────────

@router.get("/export-excel")
async def export_excel(
    is_active: Optional[bool] = Query(None, description="Filter by active status"),
    admin: Any = Depends(auth.get_admin),
    db: Any = Depends(database.get_db),
):
    """Export all products (from excel_products collection) to .xlsx."""

    filt = {}
    if is_active is not None:
        filt["is_active"] = is_active

    products = await db.excel_products.find(filt).to_list(length=10000)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products"

    # ── Header styling ──
    header_fill = PatternFill("solid", fgColor="2D6A4F")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = [
        "product_id", "product_name", "original_price", "discounted_price",
        "description", "color", "size", "product_images", "is_active",
        "stock", "category"
    ]
    col_widths = [18, 30, 15, 18, 40, 25, 20, 50, 12, 12, 20]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    # ── Data rows ──
    alt_fill = PatternFill("solid", fgColor="F0FAF4")
    for row_idx, product in enumerate(products, start=2):
        pid = product["product_id"]

        colors_docs = await db.excel_product_colors.find({"product_id": pid}).to_list(100)
        sizes_docs = await db.excel_product_sizes.find({"product_id": pid}).to_list(100)
        images_docs = await db.excel_product_images.find({"product_id": pid}).sort("sort_order", 1).to_list(100)

        colors_str = ",".join(c["color_name"] for c in colors_docs)
        sizes_str = ",".join(s["size_value"] for s in sizes_docs)
        images_str = ",".join(i["image_url"] for i in images_docs)

        row_data = [
            pid,
            product.get("product_name", ""),
            product.get("original_price", 0),
            product.get("discounted_price", ""),
            product.get("description", ""),
            colors_str,
            sizes_str,
            images_str,
            "TRUE" if product.get("is_active", True) else "FALSE",
            product.get("stock", 0),
            product.get("category", "Common"),
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))
            if fill:
                cell.fill = fill

    # Freeze header row
    ws.freeze_panes = "A2"

    # Stream response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"burhani_products_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ─────────────────────────────────────────────
#  TEMPLATE DOWNLOAD
# ─────────────────────────────────────────────

@router.get("/excel-template")
async def download_template(admin: Any = Depends(auth.get_admin)):
    """Return a pre-filled .xlsx template with example row and column descriptions."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Products Template"

    header_fill = PatternFill("solid", fgColor="1B4332")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    example_fill = PatternFill("solid", fgColor="D8F3DC")
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = [
        "product_id", "product_name", "original_price", "discounted_price",
        "description", "color", "size", "product_images", "is_active",
        "stock", "category"
    ]
    descriptions = [
        "UNIQUE ID (e.g., PROD001)", "Product display name",
        "Original price (> 0)", "Sale price (≤ original)",
        "Max 100 words", "Comma-separated (Red,Blue)",
        "Comma-separated (S,M,L,XL)", "Comma-separated image URLs",
        "TRUE or FALSE", "Quantity in stock (>= 0)",
        "Ladies Wear / Gents Wear / Kids Wear / Accessories / Common"
    ]
    col_widths = [22, 30, 18, 20, 42, 26, 22, 52, 14, 14, 28]
    example_values = [
        "PROD001", "Floral Salwar Suit", 1299, 999,
        "Beautiful floral salwar suit made from premium cotton fabric. Perfect for festive and casual occasions.",
        "Red,Blue,Black", "S,M,L,XL",
        "https://example.com/img1.jpg,https://example.com/img2.jpg",
        "TRUE", 50, "Ladies Wear"
    ]

    # Header row
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    # Description row
    desc_font = Font(italic=True, color="555555", size=9)
    for col_idx, desc in enumerate(descriptions, start=1):
        cell = ws.cell(row=2, column=col_idx, value=desc)
        cell.font = desc_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 32

    # Example data row
    for col_idx, val in enumerate(example_values, start=1):
        cell = ws.cell(row=3, column=col_idx, value=val)
        cell.fill = example_fill
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == 5))

    ws.freeze_panes = "A3"
    ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=burhani_product_template.xlsx"},
    )


# ─────────────────────────────────────────────
#  AUDIT LOGS ENDPOINT
# ─────────────────────────────────────────────

@router.get("/import-logs")
async def get_import_logs(
    limit: int = Query(20, ge=1, le=100),
    admin: Any = Depends(auth.get_admin),
    db: Any = Depends(database.get_db),
):
    """Return recent Excel import audit logs."""
    logs = (
        await db.excel_import_logs
        .find({}, {"_id": 0})
        .sort("timestamp", -1)
        .limit(limit)
        .to_list(length=limit)
    )
    # Convert datetime to ISO string for JSON
    for log in logs:
        if isinstance(log.get("timestamp"), datetime):
            log["timestamp"] = log["timestamp"].isoformat()
    return logs


# ─────────────────────────────────────────────
#  ACTIVE/INACTIVE LISTING
# ─────────────────────────────────────────────

@router.get("/excel-products")
async def list_excel_products(
    is_active: Optional[bool] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    search: Optional[str] = Query(None),
    admin: Any = Depends(auth.get_admin),
    db: Any = Depends(database.get_db),
):
    """List products managed via Excel import."""
    import math
    filt: dict = {}
    if is_active is not None:
        filt["is_active"] = is_active
    if search:
        import re as _re
        normalized_search = _re.escape(search).replace(r'\-', '[-_]').replace(r'\_', '[-_]')
        filt["$or"] = [
            {"product_name": {"$regex": search, "$options": "i"}},
            {"product_id": {"$regex": normalized_search, "$options": "i"}},
        ]

    total = await db.excel_products.count_documents(filt)
    pages = math.ceil(total / page_size) if page_size > 0 else 1
    offset = (page - 1) * page_size

    products = (
        await db.excel_products
        .find(filt, {"_id": 0})
        .skip(offset)
        .limit(page_size)
        .to_list(length=page_size)
    )

    # Populate relational data
    for p in products:
        pid = p["product_id"]
        colors = await db.excel_product_colors.find({"product_id": pid}, {"_id": 0}).to_list(50)
        sizes = await db.excel_product_sizes.find({"product_id": pid}, {"_id": 0}).to_list(50)
        images = await db.excel_product_images.find({"product_id": pid}, {"_id": 0}).sort("sort_order", 1).to_list(20)
        p["colors"] = [c["color_name"] for c in colors]
        p["sizes"] = [s["size_value"] for s in sizes]
        p["images"] = [i["image_url"] for i in images]
        if isinstance(p.get("created_at"), datetime):
            p["created_at"] = p["created_at"].isoformat()
        if isinstance(p.get("updated_at"), datetime):
            p["updated_at"] = p["updated_at"].isoformat()

    return {"items": products, "total": total, "page": page, "page_size": page_size, "pages": pages}


@router.patch("/excel-products/{product_id}/toggle")
async def toggle_product_status(
    product_id: str,
    admin: Any = Depends(auth.get_admin),
    db: Any = Depends(database.get_db),
):
    """Toggle is_active status for a single excel-imported product."""
    existing = await db.excel_products.find_one({"product_id": product_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Product not found")
    new_status = not existing.get("is_active", True)
    await db.excel_products.update_one(
        {"product_id": product_id},
        {"$set": {"is_active": new_status, "updated_at": datetime.utcnow()}}
    )
    await db.products.update_one(
        {"product_id": product_id},
        {"$set": {"is_active": new_status, "updated_at": datetime.utcnow()}}
    )
    return {"product_id": product_id, "is_active": new_status}
